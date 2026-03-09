#!/usr/bin/env python3
"""
SRD MVP App Generator
Each challenge gets a single-purpose, self-contained HTML app.
"""

import openpyxl
from pathlib import Path

XLSX = Path(
    "/Users/niff/workspace/stella/srd/sales-research/SRD_業界別課題_SaaSリサーチ.xlsx"
)
OUT = Path("/Users/niff/workspace/srd-mvp-apps/public")

# Industry color themes
COLORS = {
    1: ("#1e40af", "#3b82f6", "#dbeafe"),  # Construction - blue
    2: ("#7c2d12", "#ea580c", "#ffedd5"),  # Manufacturing - orange
    3: ("#065f46", "#10b981", "#d1fae5"),  # Logistics - green
    4: ("#7e22ce", "#a855f7", "#f3e8ff"),  # Real estate - purple
    5: ("#0e7490", "#06b6d4", "#cffafe"),  # Wholesale - cyan
    6: ("#b91c1c", "#ef4444", "#fee2e2"),  # Retail - red
    7: ("#a16207", "#eab308", "#fef9c3"),  # Hospitality - yellow
    8: ("#be185d", "#ec4899", "#fce7f3"),  # Healthcare - pink
    9: ("#4338ca", "#6366f1", "#e0e7ff"),  # Education - indigo
    10: ("#0f766e", "#14b8a6", "#ccfbf1"),  # IT - teal
    11: ("#1e3a5f", "#3b82f6", "#dbeafe"),  # Finance - navy
    12: ("#581c87", "#9333ea", "#f3e8ff"),  # Professional - violet
    13: ("#9f1239", "#f43f5e", "#ffe4e6"),  # Lifestyle - rose
    14: ("#166534", "#22c55e", "#dcfce7"),  # Agriculture - green
    15: ("#854d0e", "#f59e0b", "#fef3c7"),  # Energy - amber
    16: ("#374151", "#6b7280", "#f3f4f6"),  # Cooperative - gray
    17: ("#1e3a5f", "#2563eb", "#dbeafe"),  # Automotive - blue
    18: ("#7c2d12", "#f97316", "#fff7ed"),  # Staffing - orange
    19: ("#701a75", "#d946ef", "#fae8ff"),  # Advertising - fuchsia
    20: ("#047857", "#34d399", "#d1fae5"),  # Sports - emerald
}

# UI type classification for each challenge
UI_TYPES = {
    # dashboard, form, list, calendar, kanban, tracking, workflow, chart, checklist, alert
    "施工管理": "kanban",
    "見積": "form",
    "原価管理": "dashboard",
    "原価計算": "dashboard",
    "勤怠": "calendar",
    "出面管理": "calendar",
    "安全管理": "checklist",
    "KY活動": "checklist",
    "情報共有": "list",
    "書類作成": "form",
    "写真整理": "list",
    "台帳": "list",
    "請求": "form",
    "資格管理": "list",
    "遠隔": "tracking",
    "進捗": "kanban",
    "生産計画": "calendar",
    "在庫管理": "dashboard",
    "品質管理": "checklist",
    "設備保全": "checklist",
    "図面管理": "list",
    "技術伝承": "list",
    "受発注": "form",
    "FAX": "form",
    "サプライチェーン": "tracking",
    "日報": "form",
    "作業実績": "form",
    "配車": "calendar",
    "運転日報": "form",
    "点呼": "checklist",
    "荷待ち": "dashboard",
    "燃料": "dashboard",
    "位置情報": "tracking",
    "車両": "list",
    "労務管理": "dashboard",
    "ロケーション": "tracking",
    "車検": "alert",
    "整備": "kanban",
    "反響管理": "kanban",
    "追客": "kanban",
    "物件": "list",
    "内見": "calendar",
    "契約": "list",
    "賃貸管理": "list",
    "入金消込": "form",
    "家賃": "form",
    "オーナー報告": "form",
    "空室": "dashboard",
    "受注": "form",
    "得意先": "list",
    "商品マスタ": "list",
    "棚卸": "checklist",
    "営業": "kanban",
    "配送": "tracking",
    "与信": "dashboard",
    "返品": "form",
    "シフト管理": "calendar",
    "シフト": "calendar",
    "発注": "form",
    "POS": "dashboard",
    "情報伝達": "list",
    "顧客情報": "list",
    "顧客管理": "list",
    "EC": "dashboard",
    "売価変更": "alert",
    "廃棄": "dashboard",
    "予約": "calendar",
    "チェックイン": "form",
    "清掃": "checklist",
    "料金設定": "dashboard",
    "多言語": "form",
    "食材発注": "form",
    "原価": "dashboard",
    "レジ": "dashboard",
    "衛生管理": "checklist",
    "電子カルテ": "form",
    "レセプト": "form",
    "問診票": "form",
    "介護記録": "form",
    "介護保険": "form",
    "ケアプラン": "form",
    "スタッフ間": "list",
    "アポ帳": "calendar",
    "リコール": "alert",
    "技工物": "form",
    "生徒管理": "list",
    "月謝管理": "dashboard",
    "保護者": "list",
    "講師": "calendar",
    "入退室": "tracking",
    "研修管理": "list",
    "教材": "list",
    "受講状況": "dashboard",
    "コンプライアンス": "checklist",
    "校務": "form",
    "工数管理": "dashboard",
    "プロジェクト": "kanban",
    "PJ": "kanban",
    "アサイン": "calendar",
    "マッチング": "list",
    "ナレッジ": "list",
    "エンジニア評価": "dashboard",
    "回線開通": "kanban",
    "障害対応": "kanban",
    "問合せ": "list",
    "融資審査": "workflow",
    "稟議": "workflow",
    "AML": "checklist",
    "KYC": "checklist",
    "保険代理店": "list",
    "レガシー": "dashboard",
    "窓口業務": "form",
    "営業活動": "kanban",
    "規制報告": "form",
    "適合性": "checklist",
    "事業性評価": "dashboard",
    "案件管理": "kanban",
    "事件管理": "kanban",
    "タイムチャージ": "form",
    "顧問先": "form",
    "文書管理": "list",
    "利益相反": "checklist",
    "電子申告": "form",
    "ナレッジマネジメント": "list",
    "マーケティング": "dashboard",
    "集客": "dashboard",
    "カルテ": "form",
    "売上": "dashboard",
    "預かり品": "tracking",
    "リピート": "alert",
    "葬儀": "kanban",
    "ウェディング": "kanban",
    "キャッシュレス": "dashboard",
    "口コミ": "list",
    "圃場": "form",
    "農機": "list",
    "出荷": "form",
    "外国人": "list",
    "森林": "dashboard",
    "素材生産": "checklist",
    "養殖": "dashboard",
    "漁獲量": "form",
    "気象": "dashboard",
    "経営": "dashboard",
    "設備点検": "checklist",
    "水道管": "tracking",
    "ガス管": "tracking",
    "検針": "form",
    "CO2": "dashboard",
    "需給バランス": "dashboard",
    "料金計算": "form",
    "現場作業": "checklist",
    "技術継承": "list",
    "自治体": "kanban",
    "エネルギー": "dashboard",
    "配達": "tracking",
    "組合員": "list",
    "購買": "list",
    "共済": "list",
    "宅配": "tracking",
    "データ連携": "dashboard",
    "多能工化": "list",
    "顧客データ": "list",
    "入庫促進": "alert",
    "整備見積": "form",
    "中古車": "list",
    "商談管理": "kanban",
    "DMS": "dashboard",
    "保険・ローン": "form",
    "板金": "kanban",
    "技術情報": "list",
    "法定帳簿": "form",
    "スタッフDB": "list",
    "タイムシート": "form",
    "給与計算": "form",
    "法令遵守": "checklist",
    "開拓": "kanban",
    "応募管理": "kanban",
    "進捗管理": "kanban",
    "フォロー": "list",
    "KPI": "dashboard",
    "PJ収支": "dashboard",
    "リソース配分": "calendar",
    "レポート": "form",
    "メディアプランニング": "calendar",
    "編集・制作": "kanban",
    "校正": "form",
    "著作権": "list",
    "素材管理": "list",
    "提案書": "form",
    "販売管理": "dashboard",
    "会員管理": "list",
    "レッスン": "calendar",
    "月会費": "dashboard",
    "チケット": "form",
    "ファン": "dashboard",
    "稼働率": "dashboard",
    "イベント": "kanban",
    "物販": "dashboard",
    "インストラクター": "calendar",
    "データ分析": "dashboard",
}


def classify(challenge_name: str) -> str:
    """Classify a challenge into a UI type."""
    for keyword, ui_type in UI_TYPES.items():
        if keyword in challenge_name:
            return ui_type
    return "dashboard"  # default


def get_mock_data(ui_type: str, challenge: str, industry: str) -> str:
    """Generate mock data JS for the UI type."""
    if ui_type == "dashboard":
        return """
        const data = {
            kpi1: { label: '今月実績', value: '¥12,450,000', change: '+8.2%', up: true },
            kpi2: { label: '先月比', value: '108.2%', change: '+3.1pt', up: true },
            kpi3: { label: '未対応', value: '3件', change: '-2件', up: true },
            kpi4: { label: '完了率', value: '87.5%', change: '+5.2%', up: true },
            chartData: [65, 72, 78, 85, 82, 90, 87],
            chartLabels: ['月', '火', '水', '木', '金', '土', '日'],
            tableData: [
                ['田中工業', '¥2,300,000', '完了', '98.2%'],
                ['山本建設', '¥1,850,000', '進行中', '72.1%'],
                ['佐藤商事', '¥3,100,000', '進行中', '45.8%'],
                ['鈴木運輸', '¥980,000', '未着手', '0%'],
                ['高橋製作所', '¥1,650,000', '完了', '100%'],
            ]
        };"""
    elif ui_type == "form":
        return """
        const formFields = [
            { id: 'f1', label: '日付', type: 'date', value: new Date().toISOString().split('T')[0] },
            { id: 'f2', label: '担当者', type: 'select', options: ['田中太郎', '山本花子', '佐藤一郎', '鈴木次郎'] },
            { id: 'f3', label: '内容', type: 'textarea', placeholder: '詳細を入力...' },
            { id: 'f4', label: '金額', type: 'number', placeholder: '0' },
            { id: 'f5', label: '備考', type: 'text', placeholder: '備考があれば入力' },
        ];
        const recentEntries = [
            { date: '2026-03-08', user: '田中太郎', summary: '現場A - 完了報告', amount: '¥450,000' },
            { date: '2026-03-07', user: '山本花子', summary: '見積書作成 - B社向け', amount: '¥1,200,000' },
            { date: '2026-03-06', user: '佐藤一郎', summary: '月次レポート提出', amount: '-' },
        ];"""
    elif ui_type == "list":
        return """
        const items = [
            { id: 1, name: '田中太郎', status: 'active', detail: '営業部', date: '2026-03-01', tag: '重要' },
            { id: 2, name: '山本花子', status: 'active', detail: '技術部', date: '2026-02-28', tag: '通常' },
            { id: 3, name: '佐藤商事', status: 'pending', detail: '取引先', date: '2026-02-25', tag: '要確認' },
            { id: 4, name: '鈴木運輸', status: 'inactive', detail: '外注先', date: '2026-02-20', tag: '対応済' },
            { id: 5, name: '高橋製作所', status: 'active', detail: '協力会社', date: '2026-03-05', tag: '新規' },
            { id: 6, name: '渡辺工務店', status: 'pending', detail: '協力会社', date: '2026-03-03', tag: '要確認' },
            { id: 7, name: '伊藤設計', status: 'active', detail: '設計事務所', date: '2026-03-07', tag: '通常' },
        ];
        const filters = ['全て', '有効', '保留中', '無効'];"""
    elif ui_type == "calendar":
        return """
        const events = [
            { day: 9, title: '田中 - 早番', color: '#3b82f6', time: '06:00-15:00' },
            { day: 9, title: '山本 - 遅番', color: '#10b981', time: '14:00-23:00' },
            { day: 9, title: '佐藤 - 日勤', color: '#f59e0b', time: '09:00-18:00' },
            { day: 10, title: '田中 - 休み', color: '#ef4444', time: '-' },
            { day: 10, title: '山本 - 早番', color: '#3b82f6', time: '06:00-15:00' },
            { day: 10, title: '鈴木 - 日勤', color: '#f59e0b', time: '09:00-18:00' },
            { day: 11, title: '高橋 - 早番', color: '#3b82f6', time: '06:00-15:00' },
            { day: 11, title: '田中 - 遅番', color: '#10b981', time: '14:00-23:00' },
            { day: 12, title: '山本 - 日勤', color: '#f59e0b', time: '09:00-18:00' },
            { day: 12, title: '佐藤 - 早番', color: '#3b82f6', time: '06:00-15:00' },
        ];
        const days = Array.from({length: 31}, (_, i) => i + 1);
        const weekDays = ['日', '月', '火', '水', '木', '金', '土'];"""
    elif ui_type == "kanban":
        return """
        const columns = [
            { title: '未着手', color: '#94a3b8', items: [
                { id: 1, title: '新規案件A - 田中工業', priority: '高', assignee: '山本', date: '03/15' },
                { id: 2, title: '見積依頼 - 佐藤商事', priority: '中', assignee: '田中', date: '03/12' },
            ]},
            { title: '進行中', color: '#3b82f6', items: [
                { id: 3, title: '現場B施工 - 鈴木建設', priority: '高', assignee: '佐藤', date: '03/10' },
                { id: 4, title: '設計レビュー - 高橋設計', priority: '低', assignee: '山本', date: '03/20' },
                { id: 5, title: '資材手配 - 渡辺工務店', priority: '中', assignee: '田中', date: '03/18' },
            ]},
            { title: '確認待ち', color: '#f59e0b', items: [
                { id: 6, title: '検査報告 - 伊藤工業', priority: '高', assignee: '佐藤', date: '03/08' },
            ]},
            { title: '完了', color: '#10b981', items: [
                { id: 7, title: '引渡し - 加藤住宅', priority: '中', assignee: '山本', date: '03/05' },
                { id: 8, title: '月次報告書', priority: '低', assignee: '田中', date: '03/01' },
            ]},
        ];"""
    elif ui_type == "tracking":
        return """
        const trackingItems = [
            { id: 'TRK-001', name: '車両A (品川 300 あ 1234)', status: '配送中', location: '東京都港区', time: '10:32', progress: 75 },
            { id: 'TRK-002', name: '車両B (横浜 500 い 5678)', status: '荷積み中', location: '神奈川県横浜市', time: '10:28', progress: 25 },
            { id: 'TRK-003', name: '車両C (千葉 400 う 9012)', status: '配送完了', location: '千葉県千葉市', time: '09:45', progress: 100 },
            { id: 'TRK-004', name: '車両D (埼玉 200 え 3456)', status: '待機中', location: '埼玉県さいたま市', time: '-', progress: 0 },
            { id: 'TRK-005', name: '車両E (東京 100 お 7890)', status: '配送中', location: '東京都新宿区', time: '10:15', progress: 50 },
        ];"""
    elif ui_type == "workflow":
        return """
        const steps = [
            { name: '申請', status: 'done', user: '田中太郎', date: '2026-03-05', comment: '融資額: ¥50,000,000' },
            { name: '一次審査', status: 'done', user: '山本部長', date: '2026-03-06', comment: '財務状況良好。条件付き承認' },
            { name: '二次審査', status: 'current', user: '佐藤課長', date: '-', comment: '審査中...' },
            { name: '最終承認', status: 'pending', user: '鈴木取締役', date: '-', comment: '' },
            { name: '実行', status: 'pending', user: '経理部', date: '-', comment: '' },
        ];
        const documents = [
            { name: '融資申込書.pdf', size: '2.4MB', date: '2026-03-05' },
            { name: '決算書（3期分）.pdf', size: '8.1MB', date: '2026-03-05' },
            { name: '事業計画書.pdf', size: '1.8MB', date: '2026-03-05' },
        ];"""
    elif ui_type == "checklist":
        return """
        const checklistGroups = [
            { title: '開始前チェック', items: [
                { id: 1, text: '安全装備の着用確認', checked: true },
                { id: 2, text: '作業手順書の確認', checked: true },
                { id: 3, text: '危険箇所の確認', checked: false },
                { id: 4, text: '連絡体制の確認', checked: false },
            ]},
            { title: '作業中チェック', items: [
                { id: 5, text: '定期巡回の実施', checked: false },
                { id: 6, text: '記録写真の撮影', checked: false },
                { id: 7, text: '異常の有無確認', checked: false },
            ]},
            { title: '終了時チェック', items: [
                { id: 8, text: '作業完了の確認', checked: false },
                { id: 9, text: '片付け・清掃の確認', checked: false },
                { id: 10, text: '報告書の提出', checked: false },
            ]},
        ];"""
    elif ui_type == "alert":
        return """
        const alerts = [
            { id: 1, type: 'danger', title: '車検期限切れ間近', detail: '品川 300 あ 1234 - 残り3日（2026-03-12）', time: '2時間前' },
            { id: 2, type: 'warning', title: '在庫不足アラート', detail: '商品A（SKU-001）残り5個 - 発注点: 10個', time: '3時間前' },
            { id: 3, type: 'info', title: 'フォローアップ期限', detail: '田中工業 - 見積提出から14日経過。連絡を推奨', time: '5時間前' },
            { id: 4, type: 'danger', title: '未入金アラート', detail: '佐藤商事 - 請求書#2024-089 ¥1,500,000 期限超過7日', time: '1日前' },
            { id: 5, type: 'warning', title: '資格更新通知', detail: '山本花子 - 第一種電気工事士 更新まで30日', time: '2日前' },
            { id: 6, type: 'info', title: '定期点検予定', detail: '車両C - 12ヶ月点検 2026-03-20', time: '3日前' },
        ];"""
    return ""


def render_ui(ui_type: str, primary: str, secondary: str) -> str:
    """Return the main content HTML for each UI type."""
    if ui_type == "dashboard":
        return """
        <div class="grid grid-cols-2 md:grid-cols-4 gap-4 mb-6">
            <div class="bg-white rounded-xl p-4 shadow-sm border" id="kpi1"></div>
            <div class="bg-white rounded-xl p-4 shadow-sm border" id="kpi2"></div>
            <div class="bg-white rounded-xl p-4 shadow-sm border" id="kpi3"></div>
            <div class="bg-white rounded-xl p-4 shadow-sm border" id="kpi4"></div>
        </div>
        <div class="grid md:grid-cols-2 gap-6">
            <div class="bg-white rounded-xl p-5 shadow-sm border">
                <h3 class="font-bold text-gray-700 mb-3">推移グラフ</h3>
                <canvas id="chart" height="200"></canvas>
            </div>
            <div class="bg-white rounded-xl p-5 shadow-sm border">
                <h3 class="font-bold text-gray-700 mb-3">詳細一覧</h3>
                <table class="w-full text-sm"><thead><tr class="border-b text-left text-gray-500">
                    <th class="pb-2">名称</th><th class="pb-2">金額</th><th class="pb-2">状態</th><th class="pb-2">達成率</th>
                </tr></thead><tbody id="tableBody"></tbody></table>
            </div>
        </div>"""
    elif ui_type == "form":
        return f"""
        <div class="grid md:grid-cols-3 gap-6">
            <div class="md:col-span-2 bg-white rounded-xl p-6 shadow-sm border">
                <h3 class="font-bold text-gray-700 mb-4">新規入力</h3>
                <div id="formContainer" class="space-y-4"></div>
                <button onclick="submitForm()" class="mt-4 w-full py-2.5 rounded-lg text-white font-medium hover:opacity-90 transition" style="background:{primary}">保存する</button>
                <div id="successMsg" class="hidden mt-3 p-3 bg-green-50 text-green-700 rounded-lg text-sm text-center">保存しました</div>
            </div>
            <div class="bg-white rounded-xl p-5 shadow-sm border">
                <h3 class="font-bold text-gray-700 mb-3">最近の記録</h3>
                <div id="recentList" class="space-y-3"></div>
            </div>
        </div>"""
    elif ui_type == "list":
        return f"""
        <div class="bg-white rounded-xl shadow-sm border">
            <div class="p-4 border-b flex flex-wrap gap-2 items-center justify-between">
                <div class="flex gap-2" id="filterBtns"></div>
                <div class="flex gap-2">
                    <input type="text" placeholder="検索..." class="px-3 py-1.5 border rounded-lg text-sm" id="searchInput" oninput="filterItems()">
                    <button class="px-3 py-1.5 rounded-lg text-white text-sm" style="background:{primary}">+ 新規追加</button>
                </div>
            </div>
            <div id="listContainer" class="divide-y"></div>
        </div>"""
    elif ui_type == "calendar":
        return """
        <div class="bg-white rounded-xl shadow-sm border p-5">
            <div class="flex justify-between items-center mb-4">
                <h3 class="font-bold text-gray-700">2026年 3月</h3>
                <div class="flex gap-2">
                    <button class="px-3 py-1 border rounded text-sm">◀ 前月</button>
                    <button class="px-3 py-1 border rounded text-sm">今月</button>
                    <button class="px-3 py-1 border rounded text-sm">翌月 ▶</button>
                </div>
            </div>
            <div class="grid grid-cols-7 gap-px bg-gray-200 rounded-lg overflow-hidden" id="calendarGrid"></div>
        </div>"""
    elif ui_type == "kanban":
        return (
            """<div class="flex gap-4 overflow-x-auto pb-4" id="kanbanBoard"></div>"""
        )
    elif ui_type == "tracking":
        return """
        <div class="space-y-4" id="trackingList"></div>"""
    elif ui_type == "workflow":
        return """
        <div class="bg-white rounded-xl shadow-sm border p-6 mb-6">
            <h3 class="font-bold text-gray-700 mb-4">承認フロー</h3>
            <div id="workflowSteps" class="flex items-start gap-2 overflow-x-auto"></div>
        </div>
        <div class="bg-white rounded-xl shadow-sm border p-6">
            <h3 class="font-bold text-gray-700 mb-3">添付書類</h3>
            <div id="docList" class="space-y-2"></div>
        </div>"""
    elif ui_type == "checklist":
        return """
        <div class="space-y-4" id="checklistContainer"></div>
        <div class="mt-4 bg-white rounded-xl shadow-sm border p-4">
            <div class="flex justify-between items-center">
                <span class="text-sm text-gray-500">完了率</span>
                <span class="font-bold" id="completionRate">0%</span>
            </div>
            <div class="w-full bg-gray-200 rounded-full h-3 mt-2">
                <div class="h-3 rounded-full transition-all" id="progressBar" style="width:0%"></div>
            </div>
        </div>"""
    elif ui_type == "alert":
        return """<div class="space-y-3" id="alertList"></div>"""
    return ""


def render_script(ui_type: str, primary: str) -> str:
    """Return JS init logic for each UI type."""
    if ui_type == "dashboard":
        return f"""
        // KPIs
        ['kpi1','kpi2','kpi3','kpi4'].forEach(id => {{
            const d = data[id];
            document.getElementById(id).innerHTML = `
                <p class="text-xs text-gray-500">${{d.label}}</p>
                <p class="text-2xl font-bold mt-1">${{d.value}}</p>
                <p class="text-xs mt-1 ${{d.up?'text-green-600':'text-red-600'}}">${{d.change}}</p>`;
        }});
        // Table
        const tb = document.getElementById('tableBody');
        data.tableData.forEach(r => {{
            const sc = r[2]==='完了'?'bg-green-100 text-green-700':r[2]==='進行中'?'bg-blue-100 text-blue-700':'bg-gray-100 text-gray-600';
            tb.innerHTML += `<tr class="border-b hover:bg-gray-50"><td class="py-2">${{r[0]}}</td><td>${{r[1]}}</td><td><span class="px-2 py-0.5 rounded-full text-xs ${{sc}}">${{r[2]}}</span></td><td>${{r[3]}}</td></tr>`;
        }});
        // Chart
        new Chart(document.getElementById('chart'), {{
            type: 'line', data: {{ labels: data.chartLabels, datasets: [{{ data: data.chartData, borderColor: '{primary}', backgroundColor: '{primary}22', fill: true, tension: 0.4 }}] }},
            options: {{ plugins: {{ legend: {{ display: false }} }}, scales: {{ y: {{ beginAtZero: true }} }} }}
        }});"""
    elif ui_type == "form":
        return """
        const fc = document.getElementById('formContainer');
        formFields.forEach(f => {
            let input = '';
            if (f.type === 'textarea') input = `<textarea class="w-full border rounded-lg p-2 text-sm" rows="3" placeholder="${f.placeholder||''}"></textarea>`;
            else if (f.type === 'select') input = `<select class="w-full border rounded-lg p-2 text-sm">${f.options.map(o=>`<option>${o}</option>`).join('')}</select>`;
            else input = `<input type="${f.type}" class="w-full border rounded-lg p-2 text-sm" value="${f.value||''}" placeholder="${f.placeholder||''}">`;
            fc.innerHTML += `<div><label class="text-sm font-medium text-gray-600">${f.label}</label><div class="mt-1">${input}</div></div>`;
        });
        const rl = document.getElementById('recentList');
        recentEntries.forEach(e => {
            rl.innerHTML += `<div class="p-3 bg-gray-50 rounded-lg"><p class="text-xs text-gray-400">${e.date} | ${e.user}</p><p class="text-sm font-medium mt-1">${e.summary}</p><p class="text-xs text-gray-500 mt-1">${e.amount}</p></div>`;
        });"""
    elif ui_type == "list":
        return f"""
        const fb = document.getElementById('filterBtns');
        filters.forEach((f,i) => {{
            fb.innerHTML += `<button onclick="setFilter(this)" class="px-3 py-1 rounded-full text-sm ${{i===0?'text-white':'text-gray-600 bg-gray-100'}}" style="${{i===0?'background:{primary}':''}}">${{f}}</button>`;
        }});
        function renderItems(list) {{
            const lc = document.getElementById('listContainer');
            lc.innerHTML = '';
            list.forEach(item => {{
                const sc = item.status==='active'?'bg-green-100 text-green-700':item.status==='pending'?'bg-yellow-100 text-yellow-700':'bg-gray-100 text-gray-500';
                lc.innerHTML += `<div class="p-4 hover:bg-gray-50 flex items-center justify-between cursor-pointer">
                    <div class="flex items-center gap-3"><div class="w-10 h-10 rounded-full flex items-center justify-center text-white text-sm font-bold" style="background:{primary}">${{item.name[0]}}</div>
                    <div><p class="font-medium text-sm">${{item.name}}</p><p class="text-xs text-gray-400">${{item.detail}} | ${{item.date}}</p></div></div>
                    <div class="flex gap-2 items-center"><span class="px-2 py-0.5 rounded-full text-xs ${{sc}}">${{item.tag}}</span><span class="text-gray-300">›</span></div></div>`;
            }});
        }}
        renderItems(items);
        function filterItems() {{ const q = document.getElementById('searchInput').value.toLowerCase(); renderItems(items.filter(i => i.name.toLowerCase().includes(q) || i.detail.toLowerCase().includes(q))); }}
        function setFilter(btn) {{ document.querySelectorAll('#filterBtns button').forEach(b => {{ b.className = 'px-3 py-1 rounded-full text-sm text-gray-600 bg-gray-100'; }}); btn.className = 'px-3 py-1 rounded-full text-sm text-white'; btn.style.background = '{primary}'; }}"""
    elif ui_type == "calendar":
        return f"""
        const g = document.getElementById('calendarGrid');
        weekDays.forEach(d => {{ g.innerHTML += `<div class="bg-gray-100 p-2 text-center text-xs font-medium text-gray-500">${{d}}</div>`; }});
        // March 2026 starts on Sunday (day 0)
        for (let i = 0; i < 0; i++) g.innerHTML += `<div class="bg-white p-2 min-h-[80px]"></div>`;
        days.forEach(d => {{
            const evts = events.filter(e => e.day === d);
            const today = d === 9 ? 'ring-2 ring-offset-1' : '';
            let evtHtml = evts.map(e => `<div class="text-xs p-1 rounded mt-1 text-white truncate" style="background:${{e.color}}">${{e.title}}</div>`).join('');
            g.innerHTML += `<div class="bg-white p-2 min-h-[80px] ${{today}}" style="${{d===9?'ring-color:{primary}':''}}"><span class="text-sm ${{d===9?'font-bold':'text-gray-600'}}">${{d}}</span>${{evtHtml}}</div>`;
        }});"""
    elif ui_type == "kanban":
        return """
        const kb = document.getElementById('kanbanBoard');
        columns.forEach(col => {
            let cards = col.items.map(item => {
                const pc = item.priority==='高'?'text-red-600 bg-red-50':item.priority==='中'?'text-yellow-600 bg-yellow-50':'text-gray-500 bg-gray-50';
                return `<div class="bg-white rounded-lg p-3 shadow-sm border cursor-move hover:shadow-md transition">
                    <p class="text-sm font-medium">${item.title}</p>
                    <div class="flex justify-between items-center mt-2">
                        <span class="text-xs px-2 py-0.5 rounded ${pc}">${item.priority}</span>
                        <span class="text-xs text-gray-400">${item.assignee} | ${item.date}</span>
                    </div></div>`;
            }).join('');
            kb.innerHTML += `<div class="min-w-[280px] flex-1">
                <div class="flex items-center gap-2 mb-3"><div class="w-3 h-3 rounded-full" style="background:${col.color}"></div>
                <h3 class="font-bold text-sm text-gray-700">${col.title}</h3><span class="text-xs text-gray-400 bg-gray-100 px-2 rounded-full">${col.items.length}</span></div>
                <div class="space-y-2 bg-gray-50 rounded-xl p-3 min-h-[200px]">${cards}</div></div>`;
        });"""
    elif ui_type == "tracking":
        return f"""
        const tl = document.getElementById('trackingList');
        trackingItems.forEach(item => {{
            const sc = item.status==='配送完了'?'bg-green-100 text-green-700':item.status==='配送中'?'bg-blue-100 text-blue-700':item.status==='荷積み中'?'bg-yellow-100 text-yellow-700':'bg-gray-100 text-gray-600';
            tl.innerHTML += `<div class="bg-white rounded-xl p-4 shadow-sm border">
                <div class="flex justify-between items-start">
                    <div><p class="font-medium text-sm">${{item.name}}</p><p class="text-xs text-gray-400 mt-1">${{item.id}} | ${{item.location}} | ${{item.time}}</p></div>
                    <span class="px-2 py-1 rounded-full text-xs ${{sc}}">${{item.status}}</span>
                </div>
                <div class="mt-3"><div class="w-full bg-gray-200 rounded-full h-2"><div class="h-2 rounded-full transition-all" style="width:${{item.progress}}%;background:{primary}"></div></div>
                <p class="text-xs text-gray-400 mt-1 text-right">${{item.progress}}%</p></div></div>`;
        }});"""
    elif ui_type == "workflow":
        return f"""
        const ws = document.getElementById('workflowSteps');
        steps.forEach((s, i) => {{
            const icon = s.status==='done'?'✓':s.status==='current'?'●':'○';
            const bg = s.status==='done'?'{primary}':s.status==='current'?'#f59e0b':'#d1d5db';
            const line = i < steps.length-1 ? `<div class="flex-1 h-0.5 mt-4 mx-1" style="background:${{s.status==='done'?'{primary}':'#d1d5db'}}"></div>` : '';
            ws.innerHTML += `<div class="flex-1 min-w-[140px] text-center">
                <div class="w-8 h-8 rounded-full flex items-center justify-center text-white text-sm mx-auto" style="background:${{bg}}">${{icon}}</div>
                <p class="text-xs font-medium mt-2">${{s.name}}</p>
                <p class="text-xs text-gray-400">${{s.user}}</p>
                <p class="text-xs text-gray-400">${{s.date}}</p>
                <p class="text-xs text-gray-500 mt-1">${{s.comment}}</p></div>${{line}}`;
        }});
        const dl = document.getElementById('docList');
        documents.forEach(d => {{
            dl.innerHTML += `<div class="flex items-center justify-between p-3 bg-gray-50 rounded-lg"><div class="flex items-center gap-2"><span>📄</span><span class="text-sm">${{d.name}}</span></div><span class="text-xs text-gray-400">${{d.size}} | ${{d.date}}</span></div>`;
        }});"""
    elif ui_type == "checklist":
        return f"""
        const cc = document.getElementById('checklistContainer');
        checklistGroups.forEach(group => {{
            let items = group.items.map(item => `
                <label class="flex items-center gap-3 p-3 hover:bg-gray-50 rounded-lg cursor-pointer">
                    <input type="checkbox" ${{item.checked?'checked':''}} onchange="updateProgress()" class="w-5 h-5 rounded" style="accent-color:{primary}">
                    <span class="text-sm ${{item.checked?'line-through text-gray-400':'text-gray-700'}}">${{item.text}}</span>
                </label>`).join('');
            cc.innerHTML += `<div class="bg-white rounded-xl shadow-sm border p-4"><h3 class="font-bold text-sm text-gray-700 mb-2">${{group.title}}</h3><div class="space-y-1">${{items}}</div></div>`;
        }});
        function updateProgress() {{
            const all = document.querySelectorAll('#checklistContainer input[type=checkbox]');
            const checked = [...all].filter(c => c.checked).length;
            const pct = Math.round(checked / all.length * 100);
            document.getElementById('completionRate').textContent = pct + '%';
            document.getElementById('progressBar').style.width = pct + '%';
            document.getElementById('progressBar').style.background = '{primary}';
            all.forEach(c => {{ const span = c.nextElementSibling; if(c.checked) {{ span.classList.add('line-through','text-gray-400'); span.classList.remove('text-gray-700'); }} else {{ span.classList.remove('line-through','text-gray-400'); span.classList.add('text-gray-700'); }} }});
        }}
        updateProgress();"""
    elif ui_type == "alert":
        return f"""
        const al = document.getElementById('alertList');
        alerts.forEach(a => {{
            const styles = {{
                danger: 'border-red-200 bg-red-50',
                warning: 'border-yellow-200 bg-yellow-50',
                info: 'border-blue-200 bg-blue-50',
            }};
            const icons = {{ danger: '🔴', warning: '🟡', info: '🔵' }};
            al.innerHTML += `<div class="rounded-xl p-4 border ${{styles[a.type]}} flex items-start gap-3">
                <span class="text-lg">${{icons[a.type]}}</span>
                <div class="flex-1"><p class="font-medium text-sm">${{a.title}}</p><p class="text-xs text-gray-600 mt-1">${{a.detail}}</p></div>
                <div class="flex flex-col items-end gap-1"><span class="text-xs text-gray-400">${{a.time}}</span>
                <button class="text-xs px-2 py-1 rounded text-white" style="background:{primary}">対応</button></div></div>`;
        }});"""
    return ""


def generate_html(challenge: dict) -> str:
    ind_no = challenge["industry_no"]
    ch_no = challenge["challenge_no"]
    industry = challenge["industry"]
    ch_name = challenge["challenge"]
    pain = challenge["pain"]
    dept = challenge["dept"]
    primary, secondary, bg = COLORS.get(ind_no, ("#1e40af", "#3b82f6", "#dbeafe"))
    ui_type = classify(ch_name)
    mock_data = get_mock_data(ui_type, ch_name, industry)
    ui_html = render_ui(ui_type, primary, secondary)
    script = render_script(ui_type, primary)
    needs_chart = ui_type == "dashboard"

    return f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{ch_name} | {industry} | SRD MVP</title>
<script src="https://cdn.tailwindcss.com"></script>
{"<script src='https://cdn.jsdelivr.net/npm/chart.js'></script>" if needs_chart else ""}
<style>body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Hiragino Sans',sans-serif}}::-webkit-scrollbar{{width:6px;height:6px}}::-webkit-scrollbar-thumb{{background:#cbd5e1;border-radius:3px}}</style>
</head>
<body class="bg-gray-50 min-h-screen">
<nav class="border-b bg-white sticky top-0 z-50">
    <div class="max-w-6xl mx-auto px-4 py-3 flex items-center justify-between">
        <div class="flex items-center gap-3">
            <div class="w-8 h-8 rounded-lg flex items-center justify-center text-white text-xs font-bold" style="background:{primary}">{ind_no}</div>
            <div><p class="text-xs text-gray-400">{industry}</p><p class="font-bold text-sm">{ch_name}</p></div>
        </div>
        <div class="flex items-center gap-2">
            <span class="text-xs px-2 py-1 rounded-full" style="background:{bg};color:{primary}">{dept}</span>
            <span class="text-xs text-gray-400">MVP Demo</span>
        </div>
    </div>
</nav>
<div class="max-w-6xl mx-auto px-4 py-4">
    <div class="mb-4 p-3 rounded-lg text-sm" style="background:{bg};color:{primary}">
        <strong>課題:</strong> {pain}
    </div>
    {ui_html}
</div>
<footer class="text-center py-4 text-xs text-gray-400 mt-8">
    SRD × DX Solution Demo | {industry} #{ch_no} | Powered by Stella Group
</footer>
<script>
{mock_data}
{script}
function submitForm() {{ const m = document.getElementById('successMsg'); m.classList.remove('hidden'); setTimeout(() => m.classList.add('hidden'), 2000); }}
</script>
</body>
</html>"""


def generate_index(challenges: list) -> str:
    """Generate the index page listing all MVPs."""
    industries = {}
    for c in challenges:
        ind = c["industry"]
        if ind not in industries:
            industries[ind] = []
        industries[ind].append(c)

    cards = ""
    for ind, items in industries.items():
        ind_no = items[0]["industry_no"]
        primary, _, bg = COLORS.get(ind_no, ("#1e40af", "#3b82f6", "#dbeafe"))
        links = ""
        for item in items:
            links += f'<a href="/{ind_no}-{item["challenge_no"]}.html" class="block p-2 rounded hover:bg-gray-50 text-sm text-gray-700 truncate">{item["challenge_no"]}. {item["challenge"]}</a>'
        cards += f"""
        <div class="bg-white rounded-xl shadow-sm border overflow-hidden">
            <div class="p-4 text-white" style="background:{primary}">
                <span class="text-xs opacity-75">業界 {ind_no}</span>
                <h2 class="font-bold text-lg">{ind}</h2>
                <p class="text-xs opacity-75 mt-1">{len(items)}件のMVPデモ</p>
            </div>
            <div class="p-3 space-y-0.5 max-h-[300px] overflow-y-auto">{links}</div>
        </div>"""

    return f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>SRD 業界別課題 MVPデモ一覧</title>
<script src="https://cdn.tailwindcss.com"></script>
<style>body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Hiragino Sans',sans-serif}}</style>
</head>
<body class="bg-gray-50 min-h-screen">
<div class="max-w-6xl mx-auto px-4 py-8">
    <div class="text-center mb-8">
        <h1 class="text-3xl font-bold text-gray-800">SRD 業界別課題 MVPデモ</h1>
        <p class="text-gray-500 mt-2">20業界 × {len(challenges)}件の課題に対する単機能デモアプリ</p>
        <p class="text-xs text-gray-400 mt-1">商談時のデモツールとしてご利用ください</p>
    </div>
    <div class="grid md:grid-cols-2 lg:grid-cols-3 xl:grid-cols-4 gap-4">{cards}</div>
    <footer class="text-center py-8 text-xs text-gray-400">Powered by SRD / Stella Group</footer>
</div>
</body>
</html>"""


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["全課題一覧"]
    challenges = []
    for r in range(2, ws.max_row + 1):
        row = {
            "industry_no": ws.cell(r, 1).value,
            "industry": ws.cell(r, 2).value,
            "challenge_no": ws.cell(r, 3).value,
            "challenge": ws.cell(r, 4).value,
            "pain": ws.cell(r, 5).value,
            "dept": ws.cell(r, 6).value,
        }
        if row["challenge"]:
            challenges.append(row)

    OUT.mkdir(parents=True, exist_ok=True)

    # Generate individual pages
    for c in challenges:
        filename = f"{c['industry_no']}-{c['challenge_no']}.html"
        html = generate_html(c)
        (OUT / filename).write_text(html, encoding="utf-8")
        print(f"Generated: {filename}")

    # Generate index
    index_html = generate_index(challenges)
    (OUT / "index.html").write_text(index_html, encoding="utf-8")
    print("\nGenerated index.html")
    print(f"Total: {len(challenges)} MVP apps")


if __name__ == "__main__":
    main()
